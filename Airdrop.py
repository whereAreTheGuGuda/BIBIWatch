from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
import time
import requests
import json
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common import action_chains
from selenium.webdriver.common import keys
import openpyxl

from eth_account import Account
from eth_utils.hexadecimal import encode_hex
from eth_account.messages import encode_defunct
from xlsxwriter import Workbook

# invite_url = "https://www.peew.vip/airdrop/?r=rbad252n"
# invite_url = "https://www.peew.vip/airdrop?r=x41uitr5"
invite_url = "https://www.peew.vip/airdrop/?r=t353u04l"


class Airdrop:
    # driver: webdriver = None

    @classmethod
    def get_driver(cls):
        # if cls.driver is None:
        service = Service(
            executable_path=r"/Users/chaowang/py/matemask-inituser/chromedriver"
        )
        # metamask_home = "chrome-extension://nkbihfbeogaeaoehlefnkodbefgpgknn/home.html#new-account"
        # extron_path = r"/Users/chaowang/Library/Application Support/Google/Chrome/Default/Extensions/nmmhkkegccagdldgiimedpiccmgmieda/1.0.0.6_0"
        chrome_options = Options()
        chrome_options.add_argument("-incongnito")
        chrome_options.add_argument("--no-sandbox")
        chrome_options.add_argument("--disable-dev-shm-usage")
        chrome_options.add_argument("--disable-gpu")
        # chrome_options.add_argument("load-extension=" + extron_path)  # 这是重点
        # chrome_options.add_extension(
        # "/Users/chaowang/py/matemask-inituser/10.28.1_0.crx"
        # )
        chrome_options.add_experimental_option("detach", True)
        driver = webdriver.Chrome(options=chrome_options, service=service)
        driver.get(invite_url)
        driver.maximize_window()

        return driver

    @classmethod
    def getAllEmail(self):
        wb = openpyxl.load_workbook("email.xlsx")
        sheets = wb.sheetnames
        sheet = wb[sheets[0]]
        max_row = sheet.max_row
        emails = []
        for j in range(1, max_row + 1):
            key = sheet.cell(j, 1).value
            emails.append(key)
        print(emails)
        return emails

    @classmethod
    def createWallet(self, email):
        Account.enable_unaudited_hdwallet_features()
        account, mnemonic = Account.create_with_mnemonic()
        print(
            account.address,
            account.key.hex(),
            mnemonic,
        )
        line = "%s,%s,%s,%s" % (
            account.address,
            account.key.hex(),
            mnemonic,
            email,
        )  # mnemonic助记词
        with open(email + ".csv", "a") as f:
            f.write(line + "\n")
        return account.address

    @classmethod
    def openUrl(self, email, wallet):
        webdriver = Airdrop.get_driver()
        # webdriver.get(url=invite_url)
        time.sleep(3)
        input = webdriver.find_element(By.XPATH, '//*[@id="van-field-1-input"]')
        input.send_keys(wallet)
        email_ele = webdriver.find_element(By.XPATH, '//*[@id="van-field-2-input"]')
        email_ele.send_keys(email)
        time.sleep(2)
        # webdriver.find_element(
        #     By.XPATH, '//*[@id="__nuxt"]/div/div[1]/div[2]/div/form/button'
        # ).click()
        time.sleep(3)


if __name__ == "__main__":
    emails = Airdrop.getAllEmail()
    for i in range(len(emails)):
        if i < 270:
            continue
        email = emails[i]
        wallet = Airdrop.createWallet(email)
        time.sleep(3)
        print("准备claim wallet" + wallet + " email:" + email + " path " + str(i))
        Airdrop.openUrl(wallet=wallet, email=email)
        time.sleep(5)
